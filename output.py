from openpyxl import load_workbook
import pandas as pd


def write_to_xlsx(df):
    # The argument is a list with 12 tuples. Each tuples contains the desired data for each month.
    # It creates a xlsx with the desired format.

    # Load the existing workbook
    wb = load_workbook("SII EMITIDAS CLIENTES.xlsx")

    # Select the first worksheet
    ws = wb.worksheets[1]

    # Get the maximum row containing data
    max_row = ws.max_row

    # Delete all rows in a range
    ws.delete_rows(1, max_row)

    headers = df.columns.tolist()
    valores = df.values.tolist()

    if (
        len(headers) == 15
    ):  # No siempre se tiene la descripción de la operación, entonces a veces se inserta,
        headers.insert(
            5, "Descripción Operación"
        )  # para que luego no salte un error de tabla dinamica, y
        valores = [
            valor[:5] + [""] + valor[5:] for valor in valores
        ]  # debido a que tenemos vinculadas en el excel contra
    # ciertas columnas ciertos cuadres.
    rows = [headers] + valores
    # Write new data starting from A1
    for row in rows:
        ws.append(row)

    # Save the workbook
    wb.save("SII EMITIDAS CLIENTES.xlsx")
